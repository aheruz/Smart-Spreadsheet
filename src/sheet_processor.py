from abc import ABC
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.numbers import FORMAT_GENERAL
from typing import Any, Union, List, Dict

class SheetProcessor(ABC):
    """
    Abstract base class to process tables within a worksheet.
    """

    def __init__(self, sheet: Worksheet):
        """
        Initialize the SheetProcessor object.

        Args:
            sheet (Worksheet): The worksheet to process.
        """
        self.sheet = sheet

    def process_simple_table(self, start: Cell, end: Cell) -> List[Dict[str, Union[str, float, int]]]:
        """
        Handles a simple spreadsheet table within the specified boundary.
        Its first row is its header and the following rows are data records.
        Example:
        | Month    | Savings |
        | -------- | ------- |
        | January  | $250    |
        | February | $80     |
        | March    | $420    |

        Args:
            start (Cell): The top-left cell of the table.
            end (Cell): The bottom-right cell of the table.

        Returns:
            List[Dict[str, Union[str, float, int]]]: List of records as dictionaries.
        """
        headers = [self._format_value(cell) for cell in self.sheet[start.row][start.column-1:end.column]]
        records = []
        for row in self.sheet.iter_rows(min_row=start.row + 1, max_row=end.row, min_col=start.column, max_col=end.column):
            values = [self._format_value(cell) for cell in row]
            record = dict(zip(headers, values))
            if any(value != '' for value in values):
                records.append(self._remove_none_key_value_pairs(record))
        return records

    def process_hierarchical_table(self, start: Cell, end: Cell) -> Dict[str, Any]:
        """
        Handles a spreadsheet which has one table starting from the specified boundary.
        Its first row and first column are its headers.
        Its first column has hierarchical structural represented by the number of leading spaces.
        Some rows represents a category where the data cells are empty. Other rows represents actual data where data can be found in the data cells.
        Example:
        |                                              |30-Sep-23           |31-Oct-23           |30-Nov-23           |
        |----------------------------------------------|--------------------|--------------------|--------------------|
        |Assets                                        |                    |                    |                    |
        |   Current Assets                             |                    |                    |                    |
        |      Cash and Cash Equivalent                |                    |                    |                    |
        |         1060 TD Chequing Bank Account - #4092|587,881.66          |750,736.21          |453,234.78          |
        |         1061 TD AUD FX Currency-XXX-0283     |1,588.43            |17,457.51           |1,444.33            |
        |      Total Cash and Cash Equivalent          |$         589,470.09|$         768,193.72|$         454,679.11|
        |      1320 Prepaid Expenses                   |423,826.69          |233,127.50          |270,189.85          |
        |         1302 Prepaid License at Vid Australia|46,985.98           |68,985.98           |68,985.98           |
        |   Total Current Assets at Inc. and Australia |$      1,060,282.76 |$      1,070,307.20 |$         793,854.94|
        |Total Assets                                  |$      1,060,282.76 |$      1,070,307.20 |$         793,854.94|

        Args:
            start (Cell): The top-left cell of the table.
            end (Cell): The bottom-right cell of the table.

        Returns:
            Dict[str, Any]: Nested dictionary representing the hierarchical table.
        """
        # Extract column headers
        column_range = self.sheet[start.row][start.column:end.column]
        col_headers = [self._format_value(cell) for cell in column_range]
        # Extract row headers
        row_headers: list[str] = []
        for column in self.sheet.iter_cols(min_col=start.column, max_col=start.column, min_row=start.row + 1, max_row=end.row, values_only=False):
            row_headers = [self._format_value(cell) for cell in column]

        num_leading_space_per_level = self._calculate_num_leading_space_per_level(row_headers)
        if num_leading_space_per_level == 0:
            num_leading_space_per_level = 1

        processed_table = {}
        nodes = []

        row_iterator = self.sheet.iter_rows(
            min_row=start.row + 1,
            max_row=end.row, 
            min_col=start.column, 
            max_col=end.column, 
            values_only=False
        )
        # Process each row into the hierarchical structure
        for row in row_iterator:
            level = (len(self._format_value(row[0])) - len(self._format_value(row[0]).lstrip())) // num_leading_space_per_level
            label = self._format_value(row[0]).strip()
            data_cells = row[1:]
            nodes = nodes[:level]
            nodes.append(label)

            if any([c for c in data_cells if c.value is not None]):
                processed_table = self._add_data(processed_table, nodes, col_headers, data_cells)

        return self._remove_none_key_value_pairs(processed_table)

    @staticmethod
    def _format_value(cell) -> str:
        """
        Serialize and format the cell value to a string.

        Args:
            cell: The cell to serialize and format.

        Returns:
            str: The serialized and formatted value.
        """
        value = cell.value
        if value is None:
            return ''

        if cell.number_format == FORMAT_GENERAL:
            return f"{value}"
        else:
            return f"{value} (cell format: {cell.number_format})"

    @staticmethod
    def _remove_none_key_value_pairs(d: dict) -> dict:
        """
        Remove key-value pairs where both the key and value are None.

        Args:
            d (dict): The dictionary to clean.

        Returns:
            dict: A new dictionary with None key-value pairs removed.
        """
        return {key: value for key, value in d.items() if key or value}
    @staticmethod
    def _calculate_num_leading_space_per_level(row_headers: List[str]) -> int:
        """
        Calculate the number of leading spaces per hierarchical level.

        Args:
            row_headers (List[str]): List of row headers.

        Returns:
            int: The number of leading spaces per level.
        """
        for current_header, next_header in zip(row_headers, row_headers[1:]):
            current_spaces = len(current_header) - len(current_header.lstrip())
            next_spaces = len(next_header) - len(next_header.lstrip())
            if next_spaces != current_spaces:
                return next_spaces - current_spaces
        return 0

    def _add_data(
        self,
        processed_table: Dict[str, Any],
        nodes: List[str],
        col_headers: List[str],
        data_cells: tuple
    ) -> Dict[str, Any]:
        """
        Add data to the hierarchical table.

        Args:
            processed_table (Dict[str, Any]): The hierarchical table being built.
            nodes (List[str]): List of node labels indicating the hierarchy.
            col_headers (List[str]): Column headers.
            data_cells (tuple): Data cells corresponding to the column headers.

        Returns:
            Dict[str, Any]: Updated hierarchical table.
        """
        current_level = processed_table
        for node in nodes[:-1]:
            if node not in current_level:
                current_level[node] = {}
            current_level = current_level[node]

        current_level[nodes[-1]] = dict(zip(col_headers, [self._format_value(d) for d in data_cells]))
        return processed_table

