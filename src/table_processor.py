from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from typing import Any, Union, List, Dict, Tuple
from src.sheet_processor import SheetProcessor

class TableProcessor(SheetProcessor):
    """
    Class to identify and process tables within a worksheet.
    Inherits common methods from SheetProcessor.
    """

    def __init__(self, sheet: Worksheet):
        """
        Initialize the TableProcessor object.

        Args:
            sheet (Worksheet): The worksheet to process.
        """
        super().__init__(sheet)

    def _identify_tables(self) -> List[Tuple[Cell, Cell]]:
        """
        Identify the start and end cells of tables in the worksheet.

        Returns:
            List[Tuple[Cell, Cell]]: List of tuples containing the start and end cells of tables.
        """
        tables = []
        start_cell = None
        bottom_cell = None
        for row in self.sheet.iter_rows():
            for cell in row:
                if bottom_cell and cell.row <= bottom_cell.row and cell.column <= bottom_cell.column:
                    continue

                if not start_cell and self._is_top_left_cell(cell):
                    start_cell = cell

                if start_cell:
                    if self._is_top_right_cell(cell):
                        top_right_cell = cell
                        for bottom_row in self.sheet.iter_rows(min_row=start_cell.row, min_col=top_right_cell.column, max_col=top_right_cell.column):
                            for bottom_cell in bottom_row:
                                if self._is_bottom_right_cell(bottom_cell):
                                    bottom_right_cell = bottom_cell
                                    tables.append((start_cell, bottom_right_cell))
                                    start_cell = None
                                    break
                            if not start_cell:
                                break
        return tables

    def process_tables(self) -> List[List[Dict[str, Union[str, float, int]]]]:
        """
        Process all identified tables in the worksheet.

        Returns:
            List[List[Dict[str, Union[str, float, int]]]]: List of tables, each table is a list of records as dictionaries.
        """
        tables = []
        table_ranges = self._identify_tables()
        for start_cell, end_cell in table_ranges:
            table = self._process_table_based_on_cell_value(start_cell, end_cell)
            tables.append(table)
        return tables
    
    def _process_table_based_on_cell_value(self, start_cell: Cell, end_cell: Cell) -> Dict[str, Any]:
        """
        Processes the table based on the value of the start cell.

        Args:
            start_cell: The starting cell of the table.
            end_cell: The ending cell of the table.

        Returns:
            The processed table.
        """
        if start_cell.value is None:
            return self.process_hierarchical_table(start_cell, end_cell)
        return self.process_simple_table(start_cell, end_cell)

    def _is_header_cell(self, cell: Cell) -> bool:
        """
        Check if a cell has header style.

        Args:
            cell (Cell): The cell to check.

        Returns:
            bool: True if the cell has header style, False otherwise.
        """
        has_header_fill = (
            (cell.fill.bgColor.type == 'rgb' and cell.fill.bgColor.value == 'FF002060') or 
            (cell.fill.bgColor.type == 'indexed' and cell.fill.bgColor.value == 64)
        )
        has_borders = cell.border.top.style and (cell.border.left.style or cell.border.right.style)
        return has_header_fill and has_borders

    def _has_right_border(self, cell: Cell) -> bool:
        """
        Check if a cell has a right border.

        Args:
            cell (Cell): The cell to check.

        Returns:
            bool: True if the cell has a right border, False otherwise.
        """
        return cell.border.right.style is not None

    def _has_bottom_right_borders(self, cell: Cell) -> bool:
        """
        Check if a cell has bottom and right borders.

        Args:
            cell (Cell): The cell to check.

        Returns:
            bool: True if the cell has bottom and right borders, False otherwise.
        """
        return cell.border.bottom.style is not None and cell.border.right.style is not None

    def _is_top_left_cell(self, cell: Cell) -> bool:
        """
        Check if a cell is the top-left cell of a table.

        Args:
            cell (Cell): The cell to check.

        Returns:
            bool: True if the cell is the top-left cell of a table, False otherwise.
        """
        if self._is_header_cell(cell):
            return True
        if self._is_header_cell(cell) and \
           (cell.row-1 > 0 and not self._is_header_cell(self.sheet.cell(row=cell.row-1, column=cell.column))) and \
           (cell.column-1 > 0 and not self._is_header_cell(self.sheet.cell(row=cell.row, column=cell.column-1))):
            return True
        return False

    def _is_top_right_cell(self, cell: Cell) -> bool:
        """
        Check if a cell is the top-right cell of a table.

        Args:
            cell (Cell): The cell to check.

        Returns:
            bool: True if the cell is the top-right cell of a table, False otherwise.
        """
        if cell.value is None and not self._is_header_cell(cell):
            return False
        if cell.column == self.sheet.max_column:
            return True
        if self._is_header_cell(cell) and not self._is_header_cell(self.sheet.cell(row=cell.row, column=cell.column+1)):
            return True
        return False

    def _is_bottom_right_cell(self, cell: Cell) -> bool:
        """
        Check if a cell is the bottom-right cell of a table.

        Args:
            cell (Cell): The cell to check.

        Returns:
            bool: True if the cell is the bottom-right cell of a table, False otherwise.
        """
        if cell.value is None and not self._has_bottom_right_borders(cell):
            return False
        if cell.row == self.sheet.max_row and cell.column == self.sheet.max_column:
            return True
        if self._has_bottom_right_borders(cell) and \
           (cell.row+1 < self.sheet.max_row and not self._has_right_border(self.sheet.cell(row=cell.row+1, column=cell.column))) and \
           (cell.column+1 < self.sheet.max_column and not self._has_bottom_right_borders(self.sheet.cell(row=cell.row, column=cell.column+1))):
            return True
        return False
